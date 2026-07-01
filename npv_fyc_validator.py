"""
NPV/FYC Dashboard Validator

Scans weekly and season-to-date dashboard files and checks whether NPV and FYC
values have calculated correctly.

The validator flags:
- Excel error values such as #N/A, #DIV/0!, and #VALUE!
- Known failed-calculation fallback values
- Blank or missing NPV/FYC values
- Locked or inaccessible dashboard files
- Missing workbook tabs

Outputs:
- One Excel validation report
- Summary tab showing pass/fail status by dashboard file
- Failures Only tab showing failed checks
- Detail tab showing every checked cell

Portfolio-safe version:
All company-specific paths, brand names, workbook names, and internal references
have been removed.
"""

# ---------------------------------------------------------------------
# CONFIG
# Update these values before running the validator on a new dashboard set.
# ---------------------------------------------------------------------

ROOT_FOLDER = Path("data/raw_dashboards")
OUTPUT_DIR = Path("outputs/validation_reports")

BRANDS = {
    "Brand_A": "Brand_A",
    "Brand_B": "Brand_B",
    "Brand_C": "Brand_C",
}

SHEET_NAME = "Info"

LABEL_COL = "B"      # Column used to identify real data rows
NPV_COL = "AD"       # Column containing NPV values
FYC_COL = "AG"       # Column containing FYC values

DATA_START_ROW = 4

FAIL_VALUES = {1, 1.0}


import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter


# ---------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------

ROOT_FOLDER = Path("data/raw_dashboards")
OUTPUT_DIR = Path("outputs/validation_reports")

BRANDS = {
    "Brand_A": "Brand_A",
    "Brand_B": "Brand_B",
    "Brand_C": "Brand_C",
    "Brand_D": "Brand_D",
    "Brand_E": "Brand_E",
}

SHEET_NAME = "Info"

LABEL_COL = "B"
NPV_COL = "AD"
FYC_COL = "AG"

DATA_START_ROW = 4

FAIL_VALUES = {1, 1.0}

EXCEL_ERROR_LITERALS = {
    "#N/A",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#GETTING_DATA",
}


# ---------------------------------------------------------------------
# DATA STRUCTURE
# ---------------------------------------------------------------------

@dataclass
class Finding:
    brand: str
    dashboard_type: str
    file_name: str
    sheet: str
    row: int
    cell: str
    metric: str
    value: object
    status: str
    reason: str


# ---------------------------------------------------------------------
# FILE DISCOVERY
# ---------------------------------------------------------------------

def build_file_pattern(brand: str) -> re.Pattern:
    """
    Build a strict file-matching pattern for dashboard files.

    Expected file names:
    - Brand_A Dashboard - Week.xlsx
    - Brand_A Dashboard - STD.xlsx
    """
    pattern = rf"^{re.escape(brand)} Dashboard - (Week|STD)\.xlsx$"
    return re.compile(pattern, re.IGNORECASE)


def find_dashboard_files(root: Path, brands: dict[str, str]) -> list[dict]:
    """
    Find dashboard files for each configured brand.
    """
    found_files = []

    for folder_name, file_brand in brands.items():
        brand_folder = root / folder_name

        if not brand_folder.exists():
            print(f"[WARN] Brand folder not found: {brand_folder}")
            continue

        pattern = build_file_pattern(file_brand)

        for item in brand_folder.iterdir():
            if not item.is_file():
                continue

            if pattern.match(item.name):
                dashboard_type = "Week" if " - Week" in item.name else "STD"

                found_files.append({
                    "brand": file_brand,
                    "dashboard_type": dashboard_type,
                    "path": item,
                })

    return found_files


# ---------------------------------------------------------------------
# VALIDATION LOGIC
# ---------------------------------------------------------------------

def is_excel_error(value) -> bool:
    """
    Return True if the value is an Excel error literal.
    """
    return isinstance(value, str) and value.strip().upper() in EXCEL_ERROR_LITERALS


def is_fail_value(value) -> bool:
    """
    Return True if the value matches a known failed-calculation fallback value.
    """
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and value in FAIL_VALUES
    )


def is_blank(value) -> bool:
    """
    Return True if the value is blank or missing.
    """
    return value is None or (isinstance(value, str) and value.strip() == "")


def classify_value(value) -> tuple[str, str]:
    """
    Classify a cell value as PASS or FAIL.
    """
    if is_excel_error(value):
        return "FAIL", f"Excel error value: {value}"

    if is_fail_value(value):
        return "FAIL", f"Known failed-calculation value: {value}"

    if is_blank(value):
        return "FAIL", "Blank or missing value"

    return "PASS", ""


def check_workbook(brand: str, dashboard_type: str, path: Path) -> list[Finding]:
    """
    Check NPV and FYC columns in one dashboard workbook.
    """
    findings = []

    try:
        workbook = openpyxl.load_workbook(
            path,
            data_only=True,
            read_only=True,
        )

    except PermissionError:
        findings.append(Finding(
            brand=brand,
            dashboard_type=dashboard_type,
            file_name=path.name,
            sheet="",
            row=0,
            cell="",
            metric="FILE",
            value=None,
            status="SKIP",
            reason="File locked or in use",
        ))
        return findings

    except Exception as error:
        findings.append(Finding(
            brand=brand,
            dashboard_type=dashboard_type,
            file_name=path.name,
            sheet="",
            row=0,
            cell="",
            metric="FILE",
            value=None,
            status="SKIP",
            reason=f"Could not open file: {error}",
        ))
        return findings

    if SHEET_NAME not in workbook.sheetnames:
        findings.append(Finding(
            brand=brand,
            dashboard_type=dashboard_type,
            file_name=path.name,
            sheet=SHEET_NAME,
            row=0,
            cell="",
            metric="SHEET",
            value=None,
            status="FAIL",
            reason=f"Sheet '{SHEET_NAME}' not found",
        ))

        workbook.close()
        return findings

    sheet = workbook[SHEET_NAME]

    label_idx = column_index_from_string(LABEL_COL)
    npv_idx = column_index_from_string(NPV_COL)
    fyc_idx = column_index_from_string(FYC_COL)

    max_col_needed = max(label_idx, npv_idx, fyc_idx)

    for row_num, row in enumerate(
        sheet.iter_rows(min_row=DATA_START_ROW, max_col=max_col_needed),
        start=DATA_START_ROW,
    ):
        label_value = row[label_idx - 1].value

        if is_blank(label_value):
            continue

        metrics_to_check = [
            ("NPV", npv_idx, NPV_COL),
            ("FYC", fyc_idx, FYC_COL),
        ]

        for metric_name, col_idx, col_letter in metrics_to_check:
            value = row[col_idx - 1].value
            status, reason = classify_value(value)

            findings.append(Finding(
                brand=brand,
                dashboard_type=dashboard_type,
                file_name=path.name,
                sheet=SHEET_NAME,
                row=row_num,
                cell=f"{col_letter}{row_num}",
                metric=metric_name,
                value=value,
                status=status,
                reason=reason,
            ))

    workbook.close()
    return findings


# ---------------------------------------------------------------------
# REPORTING
# ---------------------------------------------------------------------

def build_summary_rows(all_findings: list[Finding]) -> list[list]:
    """
    Build one summary row per dashboard file.
    """
    findings_by_file = {}

    for finding in all_findings:
        key = (finding.brand, finding.dashboard_type, finding.file_name)
        findings_by_file.setdefault(key, []).append(finding)

    summary_rows = []

    for (brand, dashboard_type, file_name), findings in sorted(findings_by_file.items()):
        checked = [finding for finding in findings if finding.metric in ("NPV", "FYC")]
        failures = [finding for finding in checked if finding.status == "FAIL"]
        skipped = [finding for finding in findings if finding.status == "SKIP"]
        sheet_issues = [finding for finding in findings if finding.metric == "SHEET"]

        if skipped:
            status = "SKIPPED"
            failure_summary = skipped[0].reason

        elif sheet_issues:
            status = "FAIL"
            failure_summary = sheet_issues[0].reason

        elif failures:
            status = "FAIL"
            failure_summary = "; ".join(
                f"{failure.cell} ({failure.metric}: {failure.reason})"
                for failure in failures
            )

        else:
            status = "PASS"
            failure_summary = ""

        summary_rows.append([
            brand,
            dashboard_type,
            file_name,
            len(checked),
            len(failures),
            status,
            failure_summary,
        ])

    return summary_rows


def format_worksheet(worksheet) -> None:
    """
    Apply simple formatting to an Excel worksheet.
    """
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def write_excel_report(
    all_findings: list[Finding],
    summary_rows: list[list],
    output_dir: Path,
) -> Path:
    """
    Write validation results to one Excel report.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = output_dir / f"npv_fyc_validation_report_{timestamp}.xlsx"

    report_workbook = Workbook()

    summary_sheet = report_workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append([
        "Brand",
        "Dashboard Type",
        "File",
        "Checks Run",
        "Fail Count",
        "Status",
        "Failures",
    ])

    for row in summary_rows:
        summary_sheet.append(row)

    failures_sheet = report_workbook.create_sheet("Failures Only")
    failures_sheet.append([
        "Brand",
        "Dashboard Type",
        "File",
        "Sheet",
        "Row",
        "Cell",
        "Metric",
        "Value",
        "Status",
        "Reason",
    ])

    for finding in all_findings:
        if finding.status in ["FAIL", "SKIP"]:
            failures_sheet.append([
                finding.brand,
                finding.dashboard_type,
                finding.file_name,
                finding.sheet,
                finding.row,
                finding.cell,
                finding.metric,
                finding.value,
                finding.status,
                finding.reason,
            ])

    detail_sheet = report_workbook.create_sheet("Detail")
    detail_sheet.append([
        "Brand",
        "Dashboard Type",
        "File",
        "Sheet",
        "Row",
        "Cell",
        "Metric",
        "Value",
        "Status",
        "Reason",
    ])

    for finding in all_findings:
        detail_sheet.append([
            finding.brand,
            finding.dashboard_type,
            finding.file_name,
            finding.sheet,
            finding.row,
            finding.cell,
            finding.metric,
            finding.value,
            finding.status,
            finding.reason,
        ])

    for worksheet in report_workbook.worksheets:
        format_worksheet(worksheet)

    report_workbook.save(excel_path)

    return excel_path


# ---------------------------------------------------------------------
# RUNNER
# ---------------------------------------------------------------------

def run() -> None:
    """
    Run the full NPV/FYC validation process.
    """
    dashboard_files = find_dashboard_files(ROOT_FOLDER, BRANDS)

    if not dashboard_files:
        print("No dashboard files found. Check ROOT_FOLDER and BRANDS config.")
        return

    all_findings = []

    for dashboard_file in dashboard_files:
        all_findings.extend(check_workbook(
            brand=dashboard_file["brand"],
            dashboard_type=dashboard_file["dashboard_type"],
            path=dashboard_file["path"],
        ))

    summary_rows = build_summary_rows(all_findings)
    excel_path = write_excel_report(all_findings, summary_rows, OUTPUT_DIR)

    passed = len([row for row in summary_rows if row[5] == "PASS"])
    failed = len([row for row in summary_rows if row[5] == "FAIL"])
    skipped = len([row for row in summary_rows if row[5] == "SKIPPED"])

    print("\nDONE")
    print(f"Files scanned: {len(summary_rows)}")
    print(f"Passed: {passed}")
    print(f"Failed: {failed}")
    print(f"Skipped: {skipped}")
    print(f"Excel report: {excel_path.resolve()}")


if __name__ == "__main__":
    run()
